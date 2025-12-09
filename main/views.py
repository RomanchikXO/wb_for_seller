from typing import List
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import json
from io import BytesIO
from parsers.wildberies import get_uuid
from tasks.set_price_on_wb_from_repricer import get_marg, get_price_with_all_disc

import csv
from myapp.models import Price, Repricer, WbLk, Tags, nmids as nmids_db, Addindicators, Keywords
from django.shortcuts import render
from decorators import login_required_cust
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse
from database.DataBase import connect_to_database
from datetime import datetime, timedelta
from .services.business import business_logic_podsort, get_current_nmids, get_filter_by_articles, get_all_filters
import logging
from context_logger import ContextLogger
import docker

logger = ContextLogger(logging.getLogger("parsers"))


@login_required_cust
def main_view(request):
    data = []

    client = docker.from_env()
    containers = client.containers.list(all=True)

    for container in containers:
        data.append({
            'name': container.name,
            'id': container.short_id,
            'status': container.status,
        })
    return render(
        request,
        'main.html',
        {'data': data}
    )


@require_POST
@login_required_cust
def restart_container_view(request, container_id):
    if request.method == "POST":
        try:
            client = docker.from_env()
            container = client.containers.get(container_id)
            container.restart()
            logger.info(f"Контейнер {container_id} перезапущен пользователем {request.user}")
        except Exception as e:
            logger.error(f"Ошибка при перезапуске контейнера {container_id}: {str(e)}")
    return HttpResponseRedirect(reverse('main'))


@require_POST
@login_required_cust
def stop_container_view(request, container_id):
    if request.method == "POST":
        try:
            client = docker.from_env()
            container = client.containers.get(container_id)
            container.stop()
            logger.info(f"Контейнер {container_id} остановлен пользователем {request.user}")
        except Exception as e:
            logger.error(f"Ошибка при остановке контейнера {container_id}: {str(e)}")
    return HttpResponseRedirect(reverse('main'))


@login_required_cust
def repricer_view(request):
    page_sizes = [5, 10, 20, 50, 100]
    per_page = int(request.GET.get('per_page', 10))
    page_number = int(request.GET.get('page', 1))
    nmid_filter = request.GET.getlist('nmid')  # фильтр по артикулвм
    sort_by = request.GET.get('sort_by', '')  # значение по умолчанию
    order = request.GET.get('order', 'asc')  # asc / desc

    # Валидные поля сортировки (ключ = название в шаблоне, значение = поле в ORM)
    valid_sort_fields = {
        'redprice': 'redprice',
        'quantity': 'quantity',
        'spp': 'spp',
        'status': 'is_active',
    }

    sort_field = valid_sort_fields.get(sort_by)

    try:
        sql_query = """
            SELECT
                p.lk_id as lk_id,
                p.nmid as nmid,
                p.vendorcode as vendorcode,
                COALESCE(p.redprice, 0) as redprice,
                r.keep_price as keep_price,
                r.price_plan as price_plan,
                r.marg_or_price as marg_or_price,
                p.spp as spp,
                COALESCE(r.is_active, FALSE) AS is_active,
                COALESCE(s.total_quantity, 0) AS quantity
            FROM
                myapp_price p
            LEFT JOIN myapp_repricer r ON p.lk_id = r.lk_id AND p.nmid = r.nmid
            LEFT JOIN (
                SELECT
                    lk_id,
                    nmid,
                    SUM(quantity) AS total_quantity
                FROM
                    myapp_stocks
                GROUP BY
                    lk_id, nmid
            ) s ON p.lk_id = s.lk_id AND p.nmid = s.nmid
        """

        sql_nmid = ("SELECT p.nmid as nmid, p.vendorcode as vendorcode "
                    "FROM myapp_price p "
                    "JOIN myapp_wblk wblk "
                    "ON p.lk_id = wblk.id")
        conn = connect_to_database()
        with conn.cursor() as cursor:
            cursor.execute(sql_nmid, )
            res_nmids = cursor.fetchall()

        columns_nmids = [desc[0] for desc in cursor.description]
        nmids = [dict(zip(columns_nmids, row)) for row in res_nmids]
        combined_list = [
            {
                "nmid": item['nmid'],
                "vendorcode": item['vendorcode'],
            }
            for item in nmids
        ]
        # tail_filter_options = get_group_nmids(combined_list)
        current_ids = get_current_nmids()
        tail_filter_options = get_filter_by_articles(current_ids, size_color=True)["size_color"]

        # Добавляем фильтрацию по nmid, если она задана
        if nmid_filter:
            sql_query += " WHERE p.nmid IN (%s)" % ','.join(['%s'] * len(nmid_filter))

        # Определяем порядок сортировки
        if sort_field:
            if sort_by == 'quantity':
                sql_query += """
                        ORDER BY
                            (CASE WHEN s.total_quantity = 0 THEN 1 ELSE 0 END),
                            %s %s
                    """ % (sort_field, 'ASC' if order == 'asc' else 'DESC')
            elif sort_by == 'redprice':
                sql_query += """
                        ORDER BY
                            (CASE WHEN p.redprice IS NULL THEN 1 ELSE 0 END),
                            %s %s
                    """ % (sort_field, 'ASC' if order == 'asc' else 'DESC')
            elif sort_by == 'is_active':
                sql_query += """
                        ORDER BY
                            r.is_active %s
                    """ % ('ASC' if order == 'asc' else 'DESC')
            elif sort_by == 'spp':
                sql_query += """
                        ORDER BY
                            (CASE WHEN p.spp IS NULL THEN 1 ELSE 0 END),
                            %s %s
                    """ % (sort_field, 'ASC' if order == 'asc' else 'DESC')

        # Выполнение SQL запроса и получение данных
        conn = connect_to_database()
        with conn.cursor() as cursor:
            cursor.execute(sql_query, nmid_filter)
            rows = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]
        dict_rows = [dict(zip(columns, row)) for row in rows]

        paginator = Paginator(dict_rows, per_page)
        page_obj = paginator.get_page(page_number)


    except Exception as e:
        logger.error(f"Error in repricer_view: {e}")
        page_obj = []
        nmids = []
        paginator = None

    status_rep = Price.objects.order_by('id').values_list('main_status', flat=True).first()

    return render(request, 'repricer.html', {
        "page_obj": page_obj,
        "per_page": per_page,
        "paginator": paginator,
        "page_sizes": page_sizes,
        "nmids": combined_list,
        "nmid_filter": nmid_filter,
        "sort_by": sort_by,
        "order": order,
        "tail_filter_options": tail_filter_options,
        "status_rep": status_rep,
    })


@require_POST
@login_required_cust
def set_status_rep(request):
    payload = json.loads(request.body)

    try:
        Price.objects.all().update(main_status=payload['status'])
    except Exception as e:
        logger.error(f"Ошибка при переключении включении/отключении репрайсера в set_status_rep: {e}")
        return JsonResponse({'status': 'error'})
    return JsonResponse({'status': 'ok'})


@require_POST
@login_required_cust
def repricer_save(request):
    payload = json.loads(request.body)
    item = payload.get('item', [])

    try:
        if not item["keep_price"].isdigit(): item["keep_price"] = 0
        if not item["price_plan"].isdigit(): item["price_plan"] = 0
        lk_instance = WbLk.objects.get(id=item['lk_id'])
        Repricer.objects.update_or_create(
            lk=lk_instance,
            nmid=item['nmid'],
            defaults={
                'keep_price': item['keep_price'],
                'price_plan': item['price_plan'],
                'marg_or_price': item['marg_or_price'],
                'is_active': item['is_active']
            }
        )
    except Exception as e:
        logger.error(f"Error in repricer_save: {e}")

    return JsonResponse({'status': 'ok', 'received': len(item)})


@require_POST
@login_required_cust
def get_marg_api(request):
    payload = json.loads(request.body)
    price = payload.get('price')
    nmid = payload.get('nmid')

    sql_query = f"""
        SELECT * FROM myapp_price WHERE nmid = {int(nmid)}
    """
    conn = connect_to_database()
    with conn.cursor() as cursor:
        cursor.execute(sql_query, )
        res_nmids = cursor.fetchall()

    columns_nmids = [desc[0] for desc in cursor.description]
    nmids = [dict(zip(columns_nmids, row)) for row in res_nmids]
    nmids = nmids[0]

    try:
        price_with_disc, black_price = get_price_with_all_disc(price, nmids["spp"], nmids["discount"],
                                                               nmids["wallet_discount"])
        response = get_marg(price_with_disc, nmids["discount"], nmids["cost_price"], nmids["reject"],
                            nmids["commission"], nmids["acquiring"], nmids["nds"], nmids["usn"], nmids["drr"])
    except Exception as e:
        logger.error(f"Error in get_marg_api: {e}")
        return JsonResponse({'status': 'error', 'message': 'Ошибка в данных'})

    return JsonResponse({'status': 'ok', 'received': response})


parametrs, all_filters, warehouse_filter, turnover_change = dict(), list(), [], 0
@login_required_cust
def podsort_view(request):
    global parametrs, all_filters, warehouse_filter, turnover_change
    try:
        session_keys = ['per_page', 'period_ord', 'turnover_change', 'nmid', 'warehouse', 'alltagstb', 'sort_by', 'order',
                        'page', 'abc_filter']
        for key in session_keys:
            value = request.GET.getlist(key) if key in ['nmid', 'warehouse', 'alltagstb'] else request.GET.get(key)
            if value:
                request.session[key] = value

        export_mode = request.GET.get('export_mode', False)
        nmid_filter = request.GET.getlist('nmid', [])
        without_color_filter = request.GET.getlist('wc_filter', "") # ткань
        sizes_filter = request.GET.getlist('sz_filter', [])
        colors_filter = request.GET.getlist('cl_filter', [])
        warehouse_filter = request.GET.getlist('warehouse', "")
        alltags_filter = request.GET.getlist('alltagstb', "")
        per_page = int(request.session.get('per_page', int(request.GET.get('per_page', 10))))
        page_number = int(request.session.get('page', int(request.GET.get('page', 1))))
        sort_by = request.session.get('sort_by', request.GET.get("sort_by", ""))  # значение по умолчанию
        order = request.session.get('order', request.GET.get("order", ""))  # asc / desc
        abc_filter = request.session.get('abc_filter', request.GET.get("abc_filter", ""))
        period_ord = int(request.session.get('period_ord', int(request.GET.get('period_ord', 14))))
        turnover_change = int(request.session.get('turnover_change', int(request.GET.get('turnover_change', 40))))

        try:
            result = Addindicators.objects.filter(id=1).values_list('our_g', 'category_g').first()
            if result:
                our_g, category_g = result
            else:
                our_g, category_g = 0, 0
        except Exception as e:
            logger.error(f"Ошибка при получении Addindicators: {e}")
            our_g, category_g = 0, 0

        parametrs = {
            "export_mode": export_mode,
            "our_g": our_g,
            "category_g": category_g,
            "nmid_filter": nmid_filter,
            "without_color_filter": without_color_filter,
            "sizes_filter": sizes_filter,
            "colors_filter": colors_filter,
            "warehouse_filter": warehouse_filter,
            "alltags_filter": alltags_filter,
            "per_page": per_page,
            "page_number": page_number,
            "sort_by": sort_by,
            "order": order,
            "abc_filter": abc_filter,
            "period_ord": period_ord,
            "turnover_change": turnover_change,
        }
    except Exception as e:
        logger.error(f"Ошибка приготовления параметров {e}")
        raise

    all_filters = get_all_filters(nmid_filter, without_color_filter, sizes_filter, colors_filter)

    response = business_logic_podsort(
        warehouse_filter, parametrs,
        turnover_change, all_filters, request
    )
    return response


@login_required_cust
def autoresponse(request):
    return render(request, "autoresponse.html")


# API endpoints для автоответов
@login_required_cust
def autoresponse_status_api(request):
    """Получение и установка статуса вкл/выкл автоответов"""
    if request.method == 'GET':
        # Тестовые данные - в реальности брать из БД
        return JsonResponse({'enabled': False})
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            enabled = data.get('enabled', False)
            # Здесь должна быть логика сохранения в БД
            # Пока возвращаем тестовые данные
            logger.info(f"Статус автоответов изменен на: {enabled}")
            return JsonResponse({'status': 'ok', 'enabled': enabled})
        except Exception as e:
            logger.error(f"Ошибка обновления статуса автоответов: {e}")
            return JsonResponse({'status': 'error', 'error': str(e)}, status=400)


@login_required_cust
def autoresponse_articles_api(request):
    """Получение и обновление списка артикулов"""
    if request.method == 'GET':
        try:
            # Получаем артикулы из базы данных
            current_ids = get_current_nmids()
            if not current_ids:
                return JsonResponse({'articles': []})
            
            # Получаем vendorcode из базы
            sql_query = """
                SELECT nmid, vendorcode, use_auto_response  
                FROM myapp_nmids 
                WHERE nmid IN ({})
            """.format(', '.join(map(str, current_ids[:50])))  # Ограничиваем 50 артикулами для теста
            
            conn = connect_to_database()
            test_articles = []
            try:
                with conn.cursor() as cursor:
                    cursor.execute(sql_query)
                    rows = cursor.fetchall()
                    columns = [desc[0] for desc in cursor.description]
                    
                    for i, row in enumerate(rows):
                        row_dict = dict(zip(columns, row))
                        test_articles.append({
                            'nmid': row_dict['nmid'],
                            'vendorcode': row_dict.get('vendorcode', ''),
                            'enabled': row_dict['use_auto_response']
                        })
            finally:
                conn.close()
            
            return JsonResponse({'articles': test_articles})
        except Exception as e:
            logger.error(f"Ошибка получения артикулов: {e}")
            return JsonResponse({'status': 'error', 'error': str(e)}, status=400)
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            nmid = data.get('nmid')
            enabled = data.get('enabled', False)
            
            if not nmid:
                return JsonResponse({'status': 'error', 'error': 'Не указан nmid'}, status=400)

            nmids_db.objects.filter(nmid=nmid).update(use_auto_response=enabled)

            logger.info(f"Статус артикула {nmid} изменен на: {enabled}")
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            logger.error(f"Ошибка обновления артикула: {e}")
            return JsonResponse({'status': 'error', 'error': str(e)}, status=400)


@login_required_cust
def autoresponse_keywords_api(request):
    """CRUD операции для ключевых слов"""
    if request.method == 'GET':
        try:
            keywords = Keywords.objects.all().values('id', 'keyword', 'is_stop', 'is_positive', 'status')
            keywords_list = [
                {
                    'id': kw['id'],
                    'keyword': kw['keyword'],
                    'is_stop': kw['is_stop'],
                    'is_positive': kw['is_positive'],
                    'status': kw['status']
                }
                for kw in keywords
            ]
            return JsonResponse({'keywords': keywords_list})
        except Exception as e:
            logger.error(f"Ошибка получения ключевых слов: {e}")
            return JsonResponse({'status': 'error', 'error': str(e)}, status=400)
    
    elif request.method == 'POST':
        # Добавление нового ключевого слова или стоп слова
        try:
            data = json.loads(request.body)
            keyword_text = data.get('keyword', '').strip()
            is_stop = data.get('is_stop', False)
            
            if not keyword_text:
                return JsonResponse({'status': 'error', 'error': 'Ключевое слово не может быть пустым'}, status=400)
            
            # Проверяем, не существует ли уже такое слово
            if Keywords.objects.filter(keyword=keyword_text, is_stop=is_stop).exists():
                return JsonResponse({'status': 'error', 'error': 'Такое слово уже существует'}, status=400)
            
            keyword = Keywords.objects.create(
                keyword=keyword_text,
                is_stop=is_stop,
                is_positive=not is_stop,  # Если не стоп слово, то положительное
                status=True  # По умолчанию включено
            )
            logger.info(f"Добавлено {'стоп слово' if is_stop else 'ключевое слово'}: {keyword_text}")
            return JsonResponse({'status': 'ok', 'id': keyword.id})
        except Exception as e:
            logger.error(f"Ошибка добавления ключевого слова: {e}")
            return JsonResponse({'status': 'error', 'error': str(e)}, status=400)
    
    elif request.method == 'PUT':
        # Обновление ключевого слова (keyword или status)
        try:
            data = json.loads(request.body)
            keyword_id = data.get('id')
            
            if not keyword_id:
                return JsonResponse({'status': 'error', 'error': 'Не указан ID ключевого слова'}, status=400)
            
            try:
                keyword = Keywords.objects.get(id=keyword_id)
            except Keywords.DoesNotExist:
                return JsonResponse({'status': 'error', 'error': 'Ключевое слово не найдено'}, status=404)
            
            # Обновляем keyword, если передан
            if 'keyword' in data:
                keyword_text = data.get('keyword', '').strip()
                if not keyword_text:
                    return JsonResponse({'status': 'error', 'error': 'Ключевое слово не может быть пустым'}, status=400)
                
                # Проверяем, не существует ли уже такое слово (кроме текущего)
                if Keywords.objects.filter(keyword=keyword_text, is_stop=keyword.is_stop).exclude(id=keyword_id).exists():
                    return JsonResponse({'status': 'error', 'error': 'Такое слово уже существует'}, status=400)
                
                keyword.keyword = keyword_text
                keyword.save(update_fields=['keyword'])
                logger.info(f"Обновлено ключевое слово ID {keyword_id}: {keyword_text}")
            
            # Обновляем status, если передан
            if 'status' in data:
                keyword.status = data.get('status', False)
                keyword.save(update_fields=['status'])
                logger.info(f"Обновлен статус ключевого слова ID {keyword_id}: {keyword.status}")
            
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            logger.error(f"Ошибка обновления ключевого слова: {e}")
            return JsonResponse({'status': 'error', 'error': str(e)}, status=400)
    
    elif request.method == 'DELETE':
        # Удаление ключевого слова
        try:
            data = json.loads(request.body)
            keyword_id = data.get('id')
            
            if not keyword_id:
                return JsonResponse({'status': 'error', 'error': 'Не указан ID ключевого слова'}, status=400)
            
            try:
                keyword = Keywords.objects.get(id=keyword_id)
                keyword_text = keyword.keyword
                keyword.delete()
                logger.info(f"Удалено ключевое слово ID {keyword_id}: {keyword_text}")
                return JsonResponse({'status': 'ok'})
            except Keywords.DoesNotExist:
                return JsonResponse({'status': 'error', 'error': 'Ключевое слово не найдено'}, status=404)
        except Exception as e:
            logger.error(f"Ошибка удаления ключевого слова: {e}")
            return JsonResponse({'status': 'error', 'error': str(e)}, status=400)


@require_POST
@login_required_cust
def set_stat_nmid(request):
    try:
        data = json.loads(request.body)

        article = data.get("article")
        status = data.get("status")

        if not article or status is None:
            return JsonResponse({"error": "Не переданы обязательные параметры"}, status=400)

        updated = nmids_db.objects.filter(nmid=article).update(is_active=status)
        if updated == 0:
            return JsonResponse({"error": "Артикул не найден"}, status=404)

        return JsonResponse({'status': 'ok', 'message': "Успешно"})

    except json.JSONDecodeError:
        return JsonResponse({"error": "Некорректный JSON"}, status=400)
    except Exception as e:
        logger.error(f"Ошибка обновления статуса артикула {data if 'data' in locals() else ''}: {e}")
        return JsonResponse({"error": str(e)}, status=500)


@require_POST
@login_required_cust
def our_growth(request):
    try:
        data = json.loads(request.body)
        Addindicators.objects.update_or_create(
            id=1,
            defaults={'our_g': data['value']}
        )
        return JsonResponse({'status': 'ok', 'value': data.get('value', 0)})
    except Exception as e:
        logger.error(f"Ошибка обновления нашего роста: {e}")
        return JsonResponse({"error": str(e)}, status=400)


@require_POST
@login_required_cust
def category_growth(request):
    try:
        data = json.loads(request.body)
        Addindicators.objects.update_or_create(
            id=1,
            defaults={'category_g': data['value']}
        )
        return JsonResponse({'status': 'ok', 'value': data.get('value', 0)})
    except Exception as e:
        logger.error(f"Ошибка обновления роста категории: {e}")
        return JsonResponse({"error": str(e)}, status=400)


@require_POST
@login_required_cust
def set_tags(request):
    """
        обновляем теги у товара
    """
    try:
        data = json.loads(request.body)
        for nmid, tags in data.items():
            if tags:
                tag_ids = list(Tags.objects.filter(tag__in=tags).values_list('id', flat=True))
            else:
                tag_ids = []
            nmids_db.objects.filter(nmid=nmid).update(tag_ids=tag_ids)
    except Exception as e:
        logger.error(f"Ошибка добавления тегов к артикулу: {e}")
        return JsonResponse({"error": str(e)})
    return JsonResponse({'status': 'ok'})


@require_POST
@login_required_cust
def add_tag(request):
    try:
        data = json.loads(request.body)
        """
        тут должны добавляться новые теги в базу с тегами 
        """
        Tags.objects.create(tag=data)
    except Exception as e:
        logger.error(f"Ошибка добавления тега в БД с тегами: {e}")
        return JsonResponse({"error": str(e)})
    return JsonResponse({'status': 'ok'})


@require_POST
@login_required_cust
def export_excel(request):
    current_headers = {
        "nmid": "Артикул",
        "vendorcode": "Артикул продавца",
        "redprice": "Красная цена",
        "quantity": "Остаток",
        "spp": "spp",
        "keep_price": "Поддерживать маржу",
        "price_plan": "Поддерживать цену",
        "marg_or_price": "Маржа или Цена",
        "is_active": "Статус",

    }
    if request.method == 'POST':
        data = json.loads(request.body)
        items = data.get("items", [])

        wb = Workbook()
        ws = wb.active
        ws.title = "Repricer"

        # Заголовки (ключи словаря)
        if items:
            headers = list(items[0].keys())[1:]
            current_headers = [current_headers[i] for i in headers]
            ws.append(current_headers)

            for item in items:
                row = [item.get(col, "") if col != "marg_or_price" else ("Маржа" if item.get(col, "") else "Цена") for col in headers]
                ws.append(row)

        # Сохраняем файл в память
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=repricer_export.xlsx'
        return response


@require_POST
@login_required_cust
def upload_excel(request):
    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'error': 'Файл не получен'}, status=400)
    if not file.name.endswith('.xlsx'):
        return JsonResponse({'error': 'Допустимы только файлы .xlsx'}, status=400)

    # обработка файла: можно через openpyxl или др.
    wb = load_workbook(filename=file)
    sheet = wb.active

    try:
        data_map = {int(row[0]): int(row[1]) for row in sheet.iter_rows(values_only=True) if row[1]}
        repricer_items = Repricer.objects.filter(nmid__in=data_map.keys())

        for item in repricer_items:
            item.keep_price = data_map[item.nmid]

        Repricer.objects.bulk_update(repricer_items, ['keep_price'])

    except Exception as e:
        return JsonResponse({'error': f'Ошибка при сохранении: {e}'}, status=400)

    return JsonResponse({'status': 'ok'})


def make_data_to_load_excel(data: list) -> List[list]:
    """Подготавливаем данные для загрузки в Excel"""
    result = [
        [
            item["article"],
            item["vendorcode"],
            item["cloth"],
            item["i_color"],
            item["i_size"],
            item["orders"],
            item["stock"],
            item["ABC"],
            item["tags"],
            item["turnover_total"],
            subitem["warehouse"] if subitem else "",
            subitem["order"] if subitem else 0,
            subitem["rec_delivery"] if subitem else 0,
            subitem["stock"] if subitem else 0,
            subitem["time_available"] if subitem else 0
        ]
        for item in data
        for subitem in item.get("subitems") or [None]
    ]

    return result

@require_POST
@login_required_cust
def export_excel_podsort(request):
    """Выгрузить Excel файл из страницы подсортировщика"""
    parametrs["export_mode"] = True

    response = business_logic_podsort(
        warehouse_filter, parametrs,
        turnover_change, all_filters
    )

    # Создаём книгу и активный лист
    wb = Workbook()
    ws = wb.active
    ws.title = "Подсортировка"

    # Заголовки родительской таблицы
    headers = [
        "Артикул", "Артикул поставщика", "Ткань", "Цвет", "Размер", "Заказы", "Остатки", "АВС по размерам", "Теги",
        "Оборачиваемость общая", "Склад", "Заказы", "Рек. поставка", "Остатки", "Дней в наличии"
    ]

    row_num = 1
    header_font = Font(bold=True)

    # Пишем заголовки
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font

    # Загружаем данные
    items = response["items"].object_list

    data_to_load = make_data_to_load_excel(items)

    for item in data_to_load:
        row_num += 1
        for col_num, subitem in enumerate(item, start=1):
            ws.cell(row=row_num, column=col_num, value=subitem)

    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    response = HttpResponse(
        stream,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=podsort_export.xlsx'
    return response


@login_required_cust
def margin_view(request):
    return render(
        request,
        'margin.html',
    )

@require_POST
@login_required_cust
def get_margin_data(request):
    now_msk = datetime.now() + timedelta(hours=3)
    first_day = now_msk.replace(day=1, hour=0, minute=0, second=0, microsecond=0)


    sql_query = """
            WITH base AS (
              SELECT
                m.vendorcode,
                рисунок.value AS main_group,
                цвет.value AS color
              FROM myapp_nmids m
              LEFT JOIN LATERAL (
                SELECT (item->'value')->>0 AS value
                FROM jsonb_array_elements(m.characteristics) AS item
                WHERE (item->>'id')::int = 12
                LIMIT 1
              ) AS рисунок ON TRUE
              LEFT JOIN LATERAL (
                SELECT (item->'value')->>0 AS value
                FROM jsonb_array_elements(m.characteristics) AS item
                WHERE (item->>'id')::int = 14177449
                LIMIT 1
              ) AS цвет ON TRUE
              WHERE рисунок.value IS NOT NULL AND цвет.value IS NOT NULL
            )
            SELECT jsonb_object_agg(main_group, colors) AS result
            FROM (
              SELECT
                main_group,
                jsonb_object_agg(color, vendorcodes) AS colors
              FROM (
                SELECT
                  main_group,
                  color,
                  jsonb_agg(vendorcode) AS vendorcodes
                FROM base
                GROUP BY main_group, color
              ) AS grouped
              GROUP BY main_group
            ) AS final;
        """

    # Выполнение SQL запроса и получение данных
    conn = connect_to_database()
    with conn.cursor() as cursor:
        cursor.execute(sql_query, )
        rows = cursor.fetchall()

    columns = [desc[0] for desc in cursor.description]
    dict_rows = [dict(zip(columns, row)) for row in rows]

    items = []
    for key_1, value_1 in dict_rows[0]["result"].items():  # например key_1 "МРАМОР" value_1 "белый"
        items.append(
            {
                "id": get_uuid(),
                'name': key_1,
                'revenue_plan': '500',
                'revenue_fact': '400',
                'revenue_trend': '56',
                'margin_fact': '100',
                'drr': '50',
                'how_much_before_plan': '50',
                'revenue_plan_day': '10',
                "children": []
            }
        )
        for key_2, value_2 in value_1.items():  # например key_1 "белый" value_1 [массив артикулов]
            items[-1]["children"].append(
                {
                    "id": get_uuid(),
                    'name': key_2,
                    'revenue_plan': '500',
                    'revenue_fact': '400',
                    'revenue_trend': '56',
                    'margin_fact': '100',
                    'drr': '50',
                    'how_much_before_plan': '50',
                    'revenue_plan_day': '10',
                    "children": []
                }
            )
            for val_3 in value_2:
                items[-1]["children"][-1]["children"].append(
                    {
                        'id': get_uuid(),
                        'name': val_3,
                        'revenue_plan': '500',
                        'revenue_fact': '400',
                        'revenue_trend': '56',
                        'margin_fact': '100',
                        'drr': '50',
                        'how_much_before_plan': '50',
                        'revenue_plan_day': '10',
                    }
                )

    total = {
        'revenue_plan': '500',
        'revenue_fact': '400',
        'revenue_trend': '56',
        'margin_fact': '100',
        'drr': '50',
        'how_much_before_plan': '50',
        'revenue_plan_day': '10',
    }

    dates = [(first_day + timedelta(days=i)).strftime('%d.%m.%Y')
             for i in range((now_msk.date() - first_day.date()).days + 1)]

    return JsonResponse({
        'total': total,
        'items': items,
        'dates': dates,
    })


@login_required_cust
def shipment_view(request):
    page_sizes = [5, 10, 20, 50, 100]

    try:
        sql_query = """
            SELECT * 
            FROM myapp_shipments AS sh
            JOIN myapp_wblk AS wblk
            ON sh.lk_id = wblk.id
        """
        conn = connect_to_database()
        with conn.cursor() as cursor:
            cursor.execute(sql_query, )
            res_nmids = cursor.fetchall()

        columns_nmids = [desc[0] for desc in cursor.description]
        shipments = [dict(zip(columns_nmids, row)) for row in res_nmids]
    except Exception as e:
        logger.error(f"Ошибка получения поставок из БД в shipment_view. Ошибка: {e}")

    return render(
        request,
        'shipment.html',
        {
            "page_sizes": page_sizes,
        }
    )

@login_required_cust
def warehousewb_view(request):
    return render(
        request,
        'warehousewb.html',
    )


@require_POST
@login_required_cust
def get_warehousewb_data(request):
    try:
        sql_query = """
            SELECT 
                bw_agg.incomeid,
                bw_agg.warehousename,
                bw_agg.on_the_way,
                sp_agg.accepted,
                bw_agg.lk_name
            FROM (
                SELECT 
                    bw.incomeid,
                    bw.warehousename,
                    SUM(bw.quantity) AS on_the_way,
                    lk.name AS lk_name
                FROM myapp_betweenwarhouses AS bw
                JOIN myapp_wblk AS lk
                    ON bw.lk_id = lk.id
                GROUP BY bw.incomeid, bw.warehousename, lk.name
            ) AS bw_agg
            LEFT JOIN (
                SELECT 
                    sp."incomeId" AS incomeid,
                    SUM(sp.quantity) AS accepted
                FROM myapp_supplies AS sp
                GROUP BY sp."incomeId"
            ) AS sp_agg
            ON bw_agg.incomeid = sp_agg.incomeid
        """

        conn = connect_to_database()
        with conn.cursor() as cursor:
            cursor.execute(sql_query, )
            res_nmids = cursor.fetchall()

        columns_nmids = [desc[0] for desc in cursor.description]
        shipments = [dict(zip(columns_nmids, row)) for row in res_nmids]
    except Exception as e:
        logger.error(f"Ошибка получения поставок из БД в warehousewb_view. Ошибка: {e}")

    return JsonResponse({
        'page_obj': shipments,
    })


@require_POST
@login_required_cust
def get_warehousewb_add_data(request):
    try:
        try:
            sql_query = """
                SELECT DISTINCT name FROM myapp_wblk
            """
            conn = connect_to_database()
            with conn.cursor() as cursor:
                cursor.execute(sql_query, )
                res_nmids = cursor.fetchall()
            lks_names = [row[0] for row in res_nmids]
        except Exception as e:
            raise Exception(f"Ошибка получения ИП из БД в get_warehousewb_add_data: {e}")

        try:
            sql_query = """
                SELECT DISTINCT "warehouseName" FROM myapp_supplies
            """
            conn = connect_to_database()
            with conn.cursor() as cursor:
                cursor.execute(sql_query, )
                warehouses_data = cursor.fetchall()
            warehouses = [row[0] for row in warehouses_data]
        except Exception as e:
            raise Exception(f"Ошибка получения складов из БД в get_warehousewb_add_data: {e}")
    except Exception as e:
        logger.error(e)

    return JsonResponse({
        'lks_names': lks_names,
        'warehouses': warehouses
    })

@require_POST
@login_required_cust
def warehousewb_submit_supply(request):
    incomeid = request.POST.get('incomeid')
    warehousename = request.POST.get('warehousename')
    lk_name = request.POST.get('lk_name')

    logger.info(f"Получены данные поставки: incomeid={incomeid}, warehousename={warehousename}, lk_name={lk_name}")

    csv_file = request.FILES.get('csv_file')
    try:
        # Читаем CSV как текст (предполагаем UTF-8, можно добавить обработку кодировок)
        decoded_file = csv_file.read().decode('utf-8').splitlines()
        reader = csv.reader(decoded_file)
        rows = list(reader)

        for i, row in enumerate(rows[:10]):  # Логируем первые 10 строк, чтобы не перегружать логи
            logger.info(f"CSV строка {i + 1}: {row}")

    except Exception as e:
        logger.error(f"Ошибка чтения CSV файла: {e}")
        return JsonResponse({'error': 'Ошибка при обработке CSV файла'}, status=400)

    return JsonResponse({'status': 200,})


@csrf_exempt
def google_webhook_view(request):
    try:
        data = json.loads(request.body)
    except Exception as e:
        logger.error(f"Ошибка {e}")


# @require_POST
# @login_required_cust
# def get_dynamic_filters(request):
#     try:
#         data = json.loads(request.body)
#
#         wc_filter = set(data.get("wc_filter", "").split(",")) if data.get("wc_filter") else set()
#         if wc_filter: wc_filter = {int(i) for i in wc_filter}
#         cl_filter = set(data.get("cl_filter", "").split(",")) if data.get("cl_filter") else set()
#         if cl_filter: cl_filter = {int(i) for i in cl_filter}
#         sz_filter = set(data.get("sz_filter", "").split(",")) if data.get("sz_filter") else set()
#         if sz_filter: sz_filter = {int(i) for i in sz_filter}
#
#         filter_options_without_color = data.get("filter_options_without_color", [])
#         filter_options_colors = data.get("filter_options_colors", [])
#         filter_options_sizes = data.get("filter_options_sizes", [])
#
#         # Исходные множества
#         filtered_wc = filter_options_without_color
#         filtered_cl = filter_options_colors
#         filtered_sz = filter_options_sizes
#
#         # Если выбрано что-то, чистим остальные списки по пересечениям
#         if wc_filter: # ткань
#             nmids_allowed = wc_filter
#             if cl_filter:
#                 nmids_allowed &= cl_filter
#             if sz_filter:
#                 nmids_allowed &= sz_filter
#
#             filtered_cl = [
#                 c for c in filter_options_colors
#                 if nmids_allowed & set(c["nmids"])
#             ]
#             filtered_sz = [
#                 s for s in filter_options_sizes
#                 if nmids_allowed & set(s["nmids"])
#             ]
#
#         if cl_filter: # цвет
#             nmids_allowed = cl_filter
#             if wc_filter:
#                 nmids_allowed &= wc_filter
#             if sz_filter:
#                 nmids_allowed &= sz_filter
#
#             filtered_wc = [
#                 w for w in filter_options_without_color
#                 if nmids_allowed & set(w["nmids"])
#             ]
#             filtered_sz = [
#                 s for s in filter_options_sizes
#                 if nmids_allowed & set(s["nmids"])
#             ]
#
#         if sz_filter: # размер
#             nmids_allowed = sz_filter
#
#             if wc_filter:
#                 nmids_allowed &= wc_filter
#             if cl_filter:
#                 nmids_allowed &= cl_filter
#
#             filtered_wc = [
#                 w for w in filter_options_without_color
#                 if nmids_allowed & set(w["nmids"])
#             ]
#             filtered_cl = [
#                 c for c in filter_options_colors
#                 if nmids_allowed & set(c["nmids"])
#             ]
#     except Exception as e:
#         logger.error(f"Ошибка в get_dynamic_filters: {e}")
#         return JsonResponse({'error': e}, status=400)
#
#     return JsonResponse({
#         "filter_options_without_color": filtered_wc,
#         "filter_options_colors": filtered_cl,
#         "filter_options_sizes": filtered_sz
#     })
